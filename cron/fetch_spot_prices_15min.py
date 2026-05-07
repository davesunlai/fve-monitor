#!/usr/bin/env python3
"""
Stáhne 15min spotové ceny z OTE-CR (VDT - vnitrodenní trh) jako XLSX.
Default: dnes + zítra
Použití:
  ./fetch_spot_prices_15min.py                       # dnes + zítra
  ./fetch_spot_prices_15min.py 2026-05-07            # konkrétní den
  ./fetch_spot_prices_15min.py 2026-05-01 2026-05-07 # rozsah
"""
import sys
import io
import time
import subprocess
from datetime import date, datetime, timedelta
from urllib.request import urlopen, Request
from urllib.error import HTTPError, URLError
from openpyxl import load_workbook

UA = 'FVE-Monitor/1.0 (+https://fve.sunlai.org)'
TS = datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def http_get(url: str) -> bytes | None:
    """GET s redirect-follow + UA."""
    try:
        req = Request(url, headers={'User-Agent': UA})
        with urlopen(req, timeout=30) as resp:
            return resp.read()
    except (HTTPError, URLError) as e:
        print(f"  ✗ HTTP error: {e}")
        return None


def fetch_cnb_rate(d: date) -> float | None:
    """ČNB denní fixing pro EUR."""
    url = f'https://www.cnb.cz/cs/financni-trhy/devizovy-trh/kurzy-devizoveho-trhu/kurzy-devizoveho-trhu/denni_kurz.txt?date={d.strftime("%d.%m.%Y")}'
    body = http_get(url)
    if not body:
        return None
    for line in body.decode('utf-8', errors='ignore').split('\n'):
        parts = line.split('|')
        if len(parts) >= 5 and parts[3].strip() == 'EUR':
            try:
                amount = int(parts[2].strip())
                rate = float(parts[4].strip().replace(',', '.'))
                return rate / amount if amount > 0 else None
            except ValueError:
                return None
    return None


def fetch_ote_15min(d: date) -> list[dict] | None:
    """
    Stáhne XLSX VDT 15min pro daný den a vrátí list 96 period.
    URL: https://www.ote-cr.cz/pubweb/attachments/27/YYYY/monthMM/dayDD/VDT_15MIN_DD_MM_YYYY_CZ.xlsx
    """
    url = (f'https://www.ote-cr.cz/pubweb/attachments/27/'
           f'{d.year}/month{d.month:02d}/day{d.day:02d}/'
           f'VDT_15MIN_{d.day:02d}_{d.month:02d}_{d.year}_CZ.xlsx')

    raw = http_get(url)
    if not raw:
        return None
    if len(raw) < 1000:
        print(f"  ✗ Soubor je příliš malý ({len(raw)} B) - asi neexistuje")
        return None

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        print(f"  ✗ Chyba parsování XLSX: {e}")
        return None

    rows = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        # R7+ = data: (perioda, čas, množství, nákup, prodej, vážený průměr, min, max, poslední)
        if not row or row[0] is None:
            continue
        try:
            period = int(row[0])
            if not (1 <= period <= 96):
                continue
            time_str = str(row[1] or '').strip()
            time_from = time_str.split('-')[0] if '-' in time_str else None
            if not time_from:
                continue

            rows.append({
                'period':     period,
                'time_from':  time_from,
                'volume':     float(row[2]) if row[2] is not None else None,
                'price_avg':  float(row[5]) if row[5] is not None else None,
                'price_min':  float(row[6]) if row[6] is not None else None,
                'price_max':  float(row[7]) if row[7] is not None else None,
                'price_last': float(row[8]) if row[8] is not None else None,
            })
        except (ValueError, TypeError):
            continue

    return rows if len(rows) >= 92 else None  # 92 = tolerance pro DST den (23h × 4)


def upsert_rows(d: date, rows: list[dict], rate: float | None) -> int:
    """Hromadný UPSERT přes PHP/PDO (čte kredenciály z config.php)."""
    if not rows:
        return 0

    # Sestavíme JSON s daty a předáme do PHP scriptu, který udělá UPSERT
    import json
    payload = {
        'day': str(d),
        'rate': rate,
        'rows': [
            {
                'period': r['period'],
                'time_from': r['time_from'] + ':00',
                'volume': r['volume'],
                'price_avg': r['price_avg'],
                'price_min': r['price_min'],
                'price_max': r['price_max'],
                'price_last': r['price_last'],
            } for r in rows
        ]
    }

    php_script = """<?php
declare(strict_types=1);
require __DIR__ . '/../bootstrap.php';
use FveMonitor\\Lib\\Database;

$payload = json_decode(file_get_contents('php://stdin'), true);
if (!$payload) { fwrite(STDERR, 'invalid JSON'); exit(1); }

$pdo = Database::pdo();
$stmt = $pdo->prepare(
    'INSERT INTO spot_prices_15min
     (delivery_day, period, time_from, volume_mwh,
      price_avg_eur, price_min_eur, price_max_eur, price_last_eur,
      price_avg_czk, price_min_czk, price_max_czk, eur_czk_rate)
     VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
     ON DUPLICATE KEY UPDATE
        volume_mwh=VALUES(volume_mwh),
        price_avg_eur=VALUES(price_avg_eur),
        price_min_eur=VALUES(price_min_eur),
        price_max_eur=VALUES(price_max_eur),
        price_last_eur=VALUES(price_last_eur),
        price_avg_czk=VALUES(price_avg_czk),
        price_min_czk=VALUES(price_min_czk),
        price_max_czk=VALUES(price_max_czk),
        eur_czk_rate=VALUES(eur_czk_rate)'
);

$rate = $payload['rate'];
$day = $payload['day'];
$count = 0;
foreach ($payload['rows'] as $r) {
    $czkAvg = ($r['price_avg'] !== null && $rate) ? round($r['price_avg'] * $rate, 2) : null;
    $czkMin = ($r['price_min'] !== null && $rate) ? round($r['price_min'] * $rate, 2) : null;
    $czkMax = ($r['price_max'] !== null && $rate) ? round($r['price_max'] * $rate, 2) : null;
    $stmt->execute([
        $day, $r['period'], $r['time_from'],
        $r['volume'], $r['price_avg'], $r['price_min'], $r['price_max'], $r['price_last'],
        $czkAvg, $czkMin, $czkMax, $rate
    ]);
    $count++;
}
echo $count;
"""

    # Spustíme PHP script s JSON na stdin
    import os
    php_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_upsert_15min.php')
    if not os.path.exists(php_path):
        with open(php_path, 'w') as f:
            f.write(php_script)

    result = subprocess.run(
        ['php', php_path],
        input=json.dumps(payload), capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"  ✗ PHP chyba: {result.stderr.strip()[:300]}")
        return 0
    try:
        return int(result.stdout.strip())
    except ValueError:
        print(f"  ✗ neočekávaný výstup: {result.stdout[:100]}")
        return 0


# ────────────────────────────────────────────────
# MAIN
# ────────────────────────────────────────────────
args = sys.argv[1:]
days = []

if not args:
    days = [date.today(), date.today() + timedelta(days=1)]
elif len(args) == 1:
    days = [datetime.strptime(args[0], '%Y-%m-%d').date()]
elif len(args) == 2:
    d1 = datetime.strptime(args[0], '%Y-%m-%d').date()
    d2 = datetime.strptime(args[1], '%Y-%m-%d').date()
    while d1 <= d2:
        days.append(d1)
        d1 += timedelta(days=1)

print(f"[{TS}] Stahuji 15min ceny pro {len(days)} dnů")

ok = 0
fail = 0
rate_cache = {}

for d in days:
    print(f"→ {d}")
    rows = fetch_ote_15min(d)
    if not rows:
        fail += 1
        time.sleep(1)
        continue

    if d not in rate_cache:
        rate_cache[d] = fetch_cnb_rate(d)
    rate = rate_cache[d]

    n = upsert_rows(d, rows, rate)
    if n == 0:
        fail += 1
        continue

    eur = [r['price_avg'] for r in rows if r['price_avg'] is not None]
    if eur:
        print(f"  ✓ {n} period · min {min(eur):.2f} / Ø {sum(eur)/len(eur):.2f} / max {max(eur):.2f} EUR/MWh"
              f" · kurz {rate:.3f}" if rate else f" · kurz N/A")
    ok += 1

    if len(days) > 5:
        time.sleep(0.3)

print(f"\n[{TS}] HOTOVO: {ok} OK · {fail} FAIL")
