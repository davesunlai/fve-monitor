#!/usr/bin/env python3
"""
Stáhne DT 15min predikce z OTE-CR (denní trh, aukce) jako XLSX.
Default: zítra + pozítří (D+1, D+2)
URL: https://www.ote-cr.cz/pubweb/attachments/01/YYYY/monthMM/dayDD/DT_15MIN_DD_MM_YYYY_CZ.xlsx
Publikováno D-1 v 14:00 po skončení aukce.

Použití:
  ./fetch_spot_prices_dt15min.py                       # zítra + pozítří
  ./fetch_spot_prices_dt15min.py 2026-05-09            # konkrétní den
  ./fetch_spot_prices_dt15min.py 2026-05-01 2026-05-09 # rozsah
"""
import sys
import io
import os
import json
import time
import subprocess
from datetime import date, datetime, timedelta
from urllib.request import urlopen, Request
from urllib.error import HTTPError, URLError
from openpyxl import load_workbook

UA = 'FVE-Monitor/1.0 (+https://fve.sunlai.org)'
TS = datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def http_get(url: str) -> bytes | None:
    try:
        req = Request(url, headers={'User-Agent': UA})
        with urlopen(req, timeout=30) as resp:
            return resp.read()
    except (HTTPError, URLError) as e:
        return None


def fetch_cnb_rate(d: date) -> float | None:
    url = f'https://www.cnb.cz/cs/financni-trhy/devizovy-trh/kurzy-devizoveho-trhu/kurzy-devizoveho-trhu/denni_kurz.txt?date={d.strftime("%d.%m.%Y")}'
    body = http_get(url)
    if not body:
        return None
    for line in body.decode('utf-8', errors='ignore').split('\n'):
        parts = line.split('|')
        if len(parts) >= 5 and parts[3].strip() == 'EUR':
            try:
                amount = int(parts[2].strip())
                rate = float(parts[4].strip().replace(',', '.'))
                return rate / amount if amount > 0 else None
            except ValueError:
                return None
    return None


def fetch_ote_dt15(d: date) -> list[dict] | None:
    """Stáhne DT 15min XLSX a vrátí list 96 period."""
    url = (f'https://www.ote-cr.cz/pubweb/attachments/01/'
           f'{d.year}/month{d.month:02d}/day{d.day:02d}/'
           f'DT_15MIN_{d.day:02d}_{d.month:02d}_{d.year}_CZ.xlsx')

    raw = http_get(url)
    if not raw or len(raw) < 1000:
        print(f"  ✗ DT 15min XLSX nedostupný ({len(raw) if raw else 0} B)")
        return None

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        print(f"  ✗ Chyba parsování XLSX: {e}")
        return None

    rows = []
    # Data začínají na R24 (period 1-96), R120 = "Celkem"
    for row in ws.iter_rows(min_row=24, max_row=119, values_only=True):
        # (perioda, čas, 15min cena EUR, množství, nákup15, nákup60, prodej15, prodej60, saldo, export, import, 60min cena EUR)
        if not row or row[0] is None:
            continue
        try:
            period = int(row[0])
            if not (1 <= period <= 96):
                continue
            time_str = str(row[1] or '').strip()
            time_from = time_str.split('-')[0] if '-' in time_str else None
            if not time_from:
                continue

            def fnum(v):
                return float(v) if v is not None else None

            rows.append({
                'period':         period,
                'time_from':      time_from,
                'price_15min':    fnum(row[2]),
                'volume':         fnum(row[3]),
                'buy_15min':      fnum(row[4]),
                'buy_60min':      fnum(row[5]),
                'sell_15min':     fnum(row[6]),
                'sell_60min':     fnum(row[7]),
                'saldo':          fnum(row[8]),
                'export':         fnum(row[9]),
                'imp':            fnum(row[10]),
                'price_60min':    fnum(row[11]),
            })
        except (ValueError, TypeError):
            continue

    return rows if len(rows) >= 92 else None


def upsert_rows(d: date, rows: list[dict], rate: float | None) -> int:
    """UPSERT přes PHP/PDO helper."""
    if not rows:
        return 0

    payload = {
        'day': str(d),
        'rate': rate,
        'rows': [
            {
                'period':       r['period'],
                'time_from':    r['time_from'] + ':00',
                'price_15min':  r['price_15min'],
                'price_60min':  r['price_60min'],
                'volume':       r['volume'],
                'buy_15min':    r['buy_15min'],
                'buy_60min':    r['buy_60min'],
                'sell_15min':   r['sell_15min'],
                'sell_60min':   r['sell_60min'],
                'saldo':        r['saldo'],
                'export':       r['export'],
                'imp':          r['imp'],
            } for r in rows
        ]
    }

    php_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_upsert_dt15min.php')
    if not os.path.exists(php_path):
        php_script = '''<?php
declare(strict_types=1);
require __DIR__ . '/../bootstrap.php';
use FveMonitor\\Lib\\Database;

$payload = json_decode(file_get_contents('php://stdin'), true);
if (!$payload) { fwrite(STDERR, 'invalid JSON'); exit(1); }

$pdo = Database::pdo();
$stmt = $pdo->prepare(
    'INSERT INTO spot_prices_dt15min
     (delivery_day, period, time_from,
      price_15min_eur, price_60min_eur, volume_mwh,
      buy_15min_mwh, buy_60min_mwh, sell_15min_mwh, sell_60min_mwh,
      saldo_mwh, export_mwh, import_mwh,
      price_15min_czk, price_60min_czk, eur_czk_rate)
     VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
     ON DUPLICATE KEY UPDATE
        price_15min_eur=VALUES(price_15min_eur),
        price_60min_eur=VALUES(price_60min_eur),
        volume_mwh=VALUES(volume_mwh),
        buy_15min_mwh=VALUES(buy_15min_mwh),
        buy_60min_mwh=VALUES(buy_60min_mwh),
        sell_15min_mwh=VALUES(sell_15min_mwh),
        sell_60min_mwh=VALUES(sell_60min_mwh),
        saldo_mwh=VALUES(saldo_mwh),
        export_mwh=VALUES(export_mwh),
        import_mwh=VALUES(import_mwh),
        price_15min_czk=VALUES(price_15min_czk),
        price_60min_czk=VALUES(price_60min_czk),
        eur_czk_rate=VALUES(eur_czk_rate)'
);

$rate = $payload['rate'];
$day = $payload['day'];
$count = 0;
foreach ($payload['rows'] as $r) {
    $czk15 = ($r['price_15min'] !== null && $rate) ? round($r['price_15min'] * $rate, 2) : null;
    $czk60 = ($r['price_60min'] !== null && $rate) ? round($r['price_60min'] * $rate, 2) : null;
    $stmt->execute([
        $day, $r['period'], $r['time_from'],
        $r['price_15min'], $r['price_60min'], $r['volume'],
        $r['buy_15min'], $r['buy_60min'], $r['sell_15min'], $r['sell_60min'],
        $r['saldo'], $r['export'], $r['imp'],
        $czk15, $czk60, $rate
    ]);
    $count++;
}
echo $count;
'''
        with open(php_path, 'w') as f:
            f.write(php_script)

    result = subprocess.run(
        ['php', php_path],
        input=json.dumps(payload), capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"  ✗ PHP chyba: {result.stderr.strip()[:300]}")
        return 0
    try:
        return int(result.stdout.strip())
    except ValueError:
        print(f"  ✗ neočekávaný výstup: {result.stdout[:100]}")
        return 0


# ─── MAIN ───
args = sys.argv[1:]
days = []

if not args:
    # Default: zítra + pozítří (DT publikuje s předstihem)
    days = [date.today() + timedelta(days=1), date.today() + timedelta(days=2)]
elif len(args) == 1:
    days = [datetime.strptime(args[0], '%Y-%m-%d').date()]
elif len(args) == 2:
    d1 = datetime.strptime(args[0], '%Y-%m-%d').date()
    d2 = datetime.strptime(args[1], '%Y-%m-%d').date()
    while d1 <= d2:
        days.append(d1)
        d1 += timedelta(days=1)

print(f"[{TS}] Stahuji DT 15min predikce pro {len(days)} dnů")

ok = 0
fail = 0
rate_cache = {}

for d in days:
    print(f"→ {d}")
    rows = fetch_ote_dt15(d)
    if not rows:
        fail += 1
        time.sleep(1)
        continue

    if d not in rate_cache:
        rate_cache[d] = fetch_cnb_rate(d)
    rate = rate_cache[d]

    n = upsert_rows(d, rows, rate)
    if n == 0:
        fail += 1
        continue

    eur = [r['price_15min'] for r in rows if r['price_15min'] is not None]
    if eur:
        avg = sum(eur) / len(eur)
        print(f"  ✓ {n} period · min {min(eur):.2f} / Ø {avg:.2f} / max {max(eur):.2f} EUR/MWh"
              + (f" · kurz {rate:.3f}" if rate else " · kurz N/A"))
    ok += 1

    if len(days) > 5:
        time.sleep(0.3)

print(f"\n[{TS}] HOTOVO: {ok} OK · {fail} FAIL")
